import extract_msg
import glob
import os
import re
import pandas as pd
import hashlib
import itertools

from pandas import ExcelWriter
from openpyxl import load_workbook

# Define some variables
path = '/home/maxx/Desktop/msg/' # Input path to messages
output = 'output.xlsx' # Input path to output file
retpath = r'Return-Path: (.*)'
rec = r'Received: (.*)'
recip = r'(?:[\d]{1,3})\.(?:[\d]{1,3})\.(?:[\d]{1,3})\.(?:[\d]{1,3})'
email = r'\b[A-Za-z0-9._%-]+@[A-Za-z0-9.-]+\.[A-Za-z]{2,4}\b'
email_domain = r'[a-zA-Z0-9._%+-]+@([a-zA-Z0-9.-]+\.[a-zA-Z]{2,6})'
url = r'[a-zA-Z]+://[-a-zA-Z0-9.]+(?:/[-a-zA-Z0-9+&@#/%=~_|!:,.;]*)?(?:\?[-a-zA-Z0-9+&@#/%=~_|!:,.;]*)?'
domain = r'[a-zA-Z]+://([-a-zA-Z0-9.]+)(?:/[-a-zA-Z0-9+&@#/%=~_|!:,.;]*)?(?:\?[-a-zA-Z0-9+&@#/%=~_|!:,.;]*)?'
df_cols = ['Message_ID', 'Date', 'Sender', 'Return-Path', 'Receipt_IPs', 'To', 'CC', 'Subject', 'Body', 'Body_Emails', 'Body_URLs', 'Body_Domains']


# List dedup function
def dedup(x):
  return list(dict.fromkeys(x))

# Create destination workbook with column headers
cols = pd.DataFrame(columns = df_cols)
writer = ExcelWriter(output)
cols.to_excel(writer, sheet_name='Sheet1', index=False)
writer.save()

# Loop through messages
f = glob.glob(path + '*.msg')
for filename in f:  

    # Extract desired fields
    msg = extract_msg.Message(filename)
    msg_sender = msg.sender
    msg_to = msg.to
    msg_cc = msg.cc
    msg_date = msg.date
    msg_subj = msg.subject
    msg_head = msg.header
    msg_message = msg.body

    # Create unique message ID
    head = str(msg_head)
    head_bytes = head.encode()
    mess_hash = hashlib.sha256(head_bytes)
    mess_id = mess_hash.hexdigest()
    
    # Clean up white space in body
    msg_body = " ".join(msg_message.split())

    # Finds the return path
    retm = re.findall(retpath, head)

    # Finds reciept info and hunt IP addresses
    recm = re.findall(rec, head)
    recstr = str(recm)
    recipm = re.findall(recip, recstr)
    rec_ips = dedup(recipm)

    # String message body for search
    body = str(msg_body)

    # Hunt email addresses and dedup
    bodyemail = re.findall(email, body)
    bodyemaild = dedup(bodyemail)
    bodyedom = re.findall(email_domain, body)
    bodyedomd = dedup(bodyedom)

    # Hunt links and domains
    bodyurl = re.findall(url, body)
    bodyurld = dedup(bodyurl)
    bodydom = re.findall(domain, body)
    bodydomd = dedup(bodydom)

    # Create list of data points
    data = [mess_id, msg_date, msg_sender, retm, rec_ips, msg_to, msg_cc, msg_subj, body, bodyemaild, bodyurld, bodydomd]
    df = pd.DataFrame(data).T

    # Output to Excel
    book = load_workbook(output)
    writer = pd.ExcelWriter(output, engine='openpyxl')
    writer.book = book
    writer.sheets = {ws.title: ws for ws in book.worksheets}
    for sheetname in writer.sheets:
        df.to_excel(writer,sheet_name=sheetname, startrow=writer.sheets[sheetname].max_row, index = False,header= False)

    writer.save()
